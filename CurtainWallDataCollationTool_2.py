# coding = utf-8
# Author: Shenbq
# Date: 2021/12/15 13:22
import os
import time
from collections import Counter
import openpyxl

# 打开文件
file = 'C:\\Users\\Administrator\\Desktop\\墙板数据模板\\墙板数据模板.xlsx'
source = openpyxl.load_workbook(file)

# 获取工作表
table = source[source.sheetnames[0]]
# print(table)
# 获取行数
rows = table.max_row
# 获取列数
cols = table.max_column

data = []
# 提取数据
for i in range(rows):
    for j in range(cols):
        if table.cell(i + 1, j + 1).value != None:
            data.append(int(table.cell(i + 1, j + 1).value))

count = Counter(data).most_common()
# 升序
count = sorted(count, key=lambda x: x[0])

# 存入数据
table.cell(rows + 2, 1).value = '数值'
table.cell(rows + 2, 2).value = '个数'
for i in range(len(count)):
    table.cell(rows + 2 + i + 1, 1).value = count[i][0]
    table.cell(rows + 2 + i + 1, 2).value = count[i][1]
newDataName = '新墙板数据' + time.strftime('%Y%m%d%H%M%S', time.localtime())+ '.xls'
fullAddress = 'C:\\Users\\Administrator\\Desktop\\墙板数据模板\\整理后数据\\'+newDataName
source.save(fullAddress)
os.system(fullAddress)

