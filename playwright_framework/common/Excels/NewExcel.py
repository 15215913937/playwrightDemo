# -*- coding: utf-8 -*-
# Author: Shenbq
# Date: 2023/8/31 13:59
import os.path

import openpyxl


class Reader:
    '''
    pip install openpyxl
    '''

    def __init__(self):
        self.workbook = None
        self.sheet = None
        self.rows = 0
        self.r = 0  # 当前读取到的行数

    def open_excel(self, srcfile):
        if not os.path.isfile(srcfile):
            print("%s not exist!" % (srcfile))
            return

        openpyxl.Workbook.encoding = 'utf-8'
        self.workbook = openpyxl.load_workbook(filename=srcfile)
        self.sheet = self.workbook[self.workbook.sheetnames[0]]
        self.rows = self.sheet.max_row
        self.r = 0
        return

    def get_sheets(self):
        sheets = self.workbook.sheetnames
        return sheets

    def set_sheet(self, name):
        self.sheet = self.workbook[name]
        self.rows = self.sheet.max_row
        self.r = 0
        return

    def readline(self):
        lines = []
        for row in self.sheet.rows:
            line = []
            for cell in row:
                if cell.value is None:
                    line.append('')
                else:
                    line.append(cell.value)
            lines.append(line)
        return lines
