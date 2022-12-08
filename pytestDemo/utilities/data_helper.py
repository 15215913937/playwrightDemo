# -*- coding: utf-8 -*-
# Author: Shenbq
# Date: 2022/12/7 15:07

# 读取json文件和yaml文件
import codecs
import json
import os

import yaml
from openpyxl import load_workbook


def read_data_from_json_yaml(data_file):
    return_value = []
    # 获取data_file文件的绝对路径
    data_file_path = os.path.abspath(data_file)
    _is_yml_file = data_file_path.endswith((".yml", ".yaml"))

    with codecs.open(data_file_path, 'r', 'utf-8') as f:
        # 从yaml或者json文件中加载数据
        if _is_yml_file:
            data = yaml.safe_load(f)
        else:
            data = json.load(f)
    for i, elem in enumerate(data):
        if isinstance(data, dict):
            key, value = elem, data[elem]
            if isinstance(value, dict):
                case_data = []
                for v in value.values():
                    case_data.append(v)
                return_value.append(tuple(case_data))
            else:
                return_value.append((value,))
    return return_value


# 从excel文件中读取数据
def read_data_from_excel(excel_file, sheet_name):
    return_value = []
    if not os.path.exists(excel_file):
        raise ValueError("文件不存在！")
    wb = load_workbook(excel_file)
    for s in wb.sheetnames:
        if s == sheet_name:
            sheet = wb[sheet_name]
            for row in sheet.rows:
                return_value.append([col.value for col in row])
    return return_value
